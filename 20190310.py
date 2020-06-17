import openpyxl

# 获取工作表
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# 数据更新字典
price_update = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

# 遍历每一行
for rowNum in range(2, sheet.max_row + 1):
    # 每一行的值保存在producename变量中
    producename = sheet.cell(row=rowNum, column=1).value
    # 检查是否是需要更新的数据
    if producename in price_update:
        # 使用字典的值更新数据
        sheet.cell(row=rowNum, column=2).value = price_update[producename]
wb.save('produceSales_copy.xlsx')

