import openpyxl
from openpyxl.utils import *
from openpyxl.styles import Font

excel_name = 'MultiplicationTable.xlsx'
wb = openpyxl.Workbook(excel_name)
wb.save(excel_name)
wb = openpyxl.load_workbook(excel_name, data_only=True)
sheet = wb.active
for i in range(1, 10):
    sheet[get_column_letter(i + 1) + str(1)] = i
    sheet[get_column_letter(i + 1) + str(1)].font = Font(bold=True)
    sheet['A' + str(i + 1)] = i
    sheet['A' + str(i + 1)].font = Font(bold=True)

for i in range(1, 10):
    for j in range(1, 10):
        # x = 'A' + str(i+1)
        # y = get_column_letter(j+1) + str(1)
        sum_result = sheet['A' + str(i+1)].value * sheet[get_column_letter(j+1) + str(1)].value
        sheet[get_column_letter(j+1) + str(i+1)] = sum_result

wb.save(excel_name)
